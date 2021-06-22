#!/usr/bin/env python
# -*- coding: utf-8 -*-


import openpyxl
from util.read_config import ReadConfig
from util import get_path
from util.get_data import GetData
import os

# 读取excel测试数据
class DoExcel:

    @classmethod
    def get_data(cls, file_name):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(file_name)
        # 从配置文件读取mode的值
        mode = eval(ReadConfig().get_config(get_path.read_config_path, 'MODE', 'mode'))
        url_path = ReadConfig().get_config(get_path.read_config_path, 'ADDRESS', 'url')
        phone = getattr(GetData, 'PHONE')

        print(mode)
        test_data = []
        for key in mode:
            sheet = wb[key]
            if mode[key] == 'all':
                # 获取excel数据
                for item in range(2, sheet.max_row + 1):
                    sub_data = {}
                    sub_data['case_id'] = sheet.cell(item, 1).value
                    sub_data['title'] = sheet.cell(item, 2).value
                    sub_data['method'] = sheet.cell(item, 3).value
                    sub_data['headers'] = sheet.cell(item, 4).value
                    sub_data['url'] = url_path + sheet.cell(item, 5).value
                    if sheet.cell(item, 6).value.find('${tel}') != -1:
                        phone = phone + 1
                        setattr(GetData, 'PHONE', phone)
                        sub_data['data'] = sheet.cell(item, 6).value.replace("${tel}", str(phone))
                    else:
                        sub_data['data'] = sheet.cell(item, 6).value

                    sub_data['excepted'] = sheet.cell(item, 7).value
                    sub_data['sheet_name'] = key

                    test_data.append(sub_data)
            else:
                for case_id in mode[key]:  # 如果case_id存在的是列表里
                    sub_data = {}
                    sub_data['case_id'] = sheet.cell(case_id + 1, 1).value
                    sub_data['title'] = sheet.cell(case_id + 1, 2).value
                    sub_data['method'] = sheet.cell(case_id + 1, 3).value
                    sub_data['headers'] = sheet.cell(case_id + 1, 4).value
                    sub_data['url'] = url_path + sheet.cell(case_id + 1, 5).value
                    if sheet.cell(item, 6).value.find('${tel}') != -1:
                        phone = phone + 1
                        setattr(GetData, 'PHONE', phone)
                        sub_data['data'] = sheet.cell(case_id + 1, 6).value.replace("${tel}", str(phone))
                    else:
                        sub_data['data'] = sheet.cell(case_id + 1, 6).value

                    sub_data['excepted'] = sheet.cell(case_id + 1, 7).value
                    sub_data['sheet_name'] = key

                    test_data.append(sub_data)

        return test_data

    def write_back(self, file_name, sheet_name, row, result, testResult):
        # 打开excel，定位到表单
        wb = openpyxl.load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, 8).value = result
        sheet.cell(row, 9).value = testResult
        wb.save(file_name)


if __name__ == '__main__':
    res = DoExcel().get_data(get_path.test_excel_path)
    print(res)
